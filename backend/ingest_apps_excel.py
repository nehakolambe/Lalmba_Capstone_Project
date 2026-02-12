from __future__ import annotations

import argparse
import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import requests
from openpyxl import load_workbook
from sqlalchemy import func

try:
    from .app import create_app
    from .extensions import db
    from .models import AppDoc, AppEmbedding
except ImportError:  # Support running as a script from the backend directory.
    from app import create_app
    from extensions import db
    from models import AppDoc, AppEmbedding


logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"), format="%(levelname)s %(message)s")
logger = logging.getLogger("ingest_apps_excel")

OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434").rstrip("/")
OLLAMA_EMBED_MODEL = os.getenv("OLLAMA_EMBED_MODEL", "embeddinggemma")
OLLAMA_EMBED_URL = f"{OLLAMA_BASE_URL}/api/embed"
OLLAMA_EMBED_FALLBACK_URL = f"{OLLAMA_BASE_URL}/api/embeddings"


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _parse_bool(value: object) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False
    text = str(value).strip().lower()
    if not text:
        return None
    if text in {"yes", "y", "true", "1", "ok", "available", "supported"}:
        return True
    if text in {"no", "n", "false", "0", "not", "none", "unsupported"}:
        return False
    return None


def _normalize_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _normalize_keep_installed(value: object) -> Optional[str]:
    parsed = _parse_bool(value)
    if parsed is True:
        return "Yes"
    if parsed is False:
        return "No"
    return _normalize_text(value)


def _map_header(header: object) -> Optional[str]:
    key = _normalize_header(header)
    if not key:
        return None
    exact = {
        "appname": "app_name",
        "application": "app_name",
        "applications": "app_name",
        "name": "app_name",
        "category": "category",
        "categories": "category",
        "requiresinternet": "requires_internet",
        "internetrequired": "requires_internet",
        "needsinternet": "requires_internet",
        "cost": "cost",
        "price": "cost",
        "swahili": "swahili_support",
        "kiswahili": "swahili_support",
        "keptinstalled": "keep_installed",
        "keepinstalled": "keep_installed",
        "keep": "keep_installed",
        "impact": "impact",
        "description": "description",
        "summary": "description",
    }
    if key in exact:
        return exact[key]
    if "app" in key and "name" in key:
        return "app_name"
    if "category" in key:
        return "category"
    if "requires" in key and "internet" in key:
        return "requires_internet"
    if "offline" in key:
        return "offline"
    if "swahili" in key or "kiswahili" in key:
        return "swahili_support"
    if "keep" in key and ("install" in key or "installed" in key):
        return "keep_installed"
    if "impact" in key:
        return "impact"
    if "description" in key or "summary" in key or "details" in key:
        return "description"
    if "cost" in key or "price" in key:
        return "cost"
    return None


def _build_doc_text(row: Dict[str, object]) -> str:
    app_name = row.get("app_name") or "Unknown"
    category = row.get("category") or "Unknown"
    requires_internet = row.get("requires_internet")
    offline = None if requires_internet is None else not requires_internet
    swahili = row.get("swahili_support")
    keep_installed = row.get("keep_installed") or "Unknown"
    impact = row.get("impact") or "Unknown"
    description = row.get("description") or "Not provided"
    offline_text = "Unknown" if offline is None else ("Yes" if offline else "No")
    swahili_text = "Unknown" if swahili is None else ("Yes" if swahili else "No")
    return (
        f"App: {app_name}. Category: {category}. Offline: {offline_text}. "
        f"Swahili: {swahili_text}. Keep: {keep_installed}. "
        f"Impact: {impact}. Summary: {description}."
    )


def _extract_embedding(data: object) -> Optional[List[float]]:
    if not isinstance(data, dict):
        return None
    embedding = data.get("embedding")
    if isinstance(embedding, list):
        return embedding
    embeddings = data.get("embeddings")
    if isinstance(embeddings, list) and embeddings:
        first = embeddings[0]
        if isinstance(first, list):
            return first
    return None


def _embed_text(text: str, *, model: str, base_url: str) -> List[float]:
    if not text.strip():
        return []

    embed_url = f"{base_url.rstrip('/')}/api/embed"
    fallback_url = f"{base_url.rstrip('/')}/api/embeddings"
    payload = {"model": model, "input": text}
    try:
        response = requests.post(embed_url, json=payload, timeout=60)
    except requests.RequestException as exc:
        raise RuntimeError(f"Could not reach Ollama embeddings: {exc}") from exc

    if response.status_code == 404:
        try:
            response = requests.post(fallback_url, json={"model": model, "prompt": text}, timeout=60)
        except requests.RequestException as exc:
            raise RuntimeError(f"Could not reach Ollama embeddings: {exc}") from exc

    try:
        data = response.json()
    except ValueError as exc:
        raise RuntimeError("Unexpected response from Ollama embeddings") from exc

    if not response.ok:
        reason = data.get("error") if isinstance(data, dict) else response.text
        raise RuntimeError(f"Ollama embedding error: {reason}")

    embedding = _extract_embedding(data)
    if not embedding:
        raise RuntimeError("Ollama embedding returned empty result")
    return embedding


def _resolve_excel_path(raw_path: Optional[str]) -> Path:
    if raw_path:
        return Path(raw_path).expanduser()

    candidates = [
        Path(__file__).resolve().parents[1] / "Endless OS Applications.xlsx",
        Path(__file__).resolve().parents[2] / "Endless OS Applications.xlsx",
        Path.home() / "Downloads" / "Endless OS Applications.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("Could not find 'Endless OS Applications.xlsx'. Provide --file path.")


def _load_rows(path: Path, sheet_name: Optional[str]) -> List[Dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    header_map: Dict[int, str] = {}
    for idx, col in enumerate(header):
        mapped = _map_header(col)
        if mapped:
            header_map[idx] = mapped

    normalized_rows: List[Dict[str, object]] = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        data: Dict[str, object] = {"offline": None}
        for idx, cell in enumerate(row):
            key = header_map.get(idx)
            if not key:
                continue
            if key == "app_name":
                data["app_name"] = _normalize_text(cell)
            elif key == "category":
                data["category"] = _normalize_text(cell)
            elif key == "requires_internet":
                data["requires_internet"] = _parse_bool(cell)
            elif key == "offline":
                data["offline"] = _parse_bool(cell)
            elif key == "cost":
                data["cost"] = _normalize_text(cell)
            elif key == "swahili_support":
                data["swahili_support"] = _parse_bool(cell)
            elif key == "keep_installed":
                data["keep_installed"] = _normalize_keep_installed(cell)
            elif key == "impact":
                data["impact"] = _normalize_text(cell)
            elif key == "description":
                data["description"] = _normalize_text(cell)

        if not data.get("app_name"):
            continue

        if data.get("requires_internet") is None and data.get("offline") is not None:
            data["requires_internet"] = not bool(data["offline"])

        data.pop("offline", None)
        normalized_rows.append(data)

    return normalized_rows


def ingest_excel(path: Path, *, sheet_name: Optional[str], embed_model: str, base_url: str) -> int:
    rows = _load_rows(path, sheet_name)
    if not rows:
        logger.warning("No rows found in %s", path)
        return 0

    source_name = path.name
    created = 0
    for row in rows:
        app_name = row.get("app_name")
        if not app_name:
            continue

        app = (
            db.session.query(AppDoc)
            .filter(func.lower(AppDoc.app_name) == str(app_name).lower())
            .first()
        )
        if app is None:
            app = AppDoc(app_name=app_name, created_at=datetime.utcnow())
            created += 1
            db.session.add(app)

        app.category = row.get("category")
        app.requires_internet = row.get("requires_internet")
        app.cost = row.get("cost")
        app.swahili_support = row.get("swahili_support")
        app.keep_installed = row.get("keep_installed")
        app.impact = row.get("impact")
        app.description = row.get("description")
        app.source = source_name

        db.session.flush()
        doc_text = _build_doc_text(
            {
                "app_name": app.app_name,
                "category": app.category,
                "requires_internet": app.requires_internet,
                "swahili_support": app.swahili_support,
                "keep_installed": app.keep_installed,
                "impact": app.impact,
                "description": app.description,
            }
        )
        embedding = _embed_text(doc_text, model=embed_model, base_url=base_url)

        existing_embedding = (
            db.session.query(AppEmbedding)
            .filter(AppEmbedding.app_doc_id == app.id, AppEmbedding.embedding_model == embed_model)
            .first()
        )
        if existing_embedding is None:
            existing_embedding = AppEmbedding(
                app_doc_id=app.id,
                embedding_model=embed_model,
                created_at=datetime.utcnow(),
            )
            db.session.add(existing_embedding)

        existing_embedding.embedding_json = json.dumps(embedding)
        existing_embedding.created_at = datetime.utcnow()

    db.session.commit()
    return created


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest Endless OS applications from Excel into SQLite.")
    parser.add_argument("--file", help="Path to 'Endless OS Applications.xlsx'")
    parser.add_argument("--sheet", help="Optional sheet name")
    parser.add_argument("--embed-model", default=OLLAMA_EMBED_MODEL, help="Ollama embedding model name")
    parser.add_argument("--base-url", default=OLLAMA_BASE_URL, help="Ollama base URL")
    args = parser.parse_args()

    path = _resolve_excel_path(args.file)
    if not path.exists():
        raise FileNotFoundError(path)

    app = create_app()
    with app.app_context():
        created = ingest_excel(
            path,
            sheet_name=args.sheet,
            embed_model=args.embed_model,
            base_url=args.base_url,
        )
    logger.info("Ingestion complete. New apps added: %s", created)


if __name__ == "__main__":
    main()
